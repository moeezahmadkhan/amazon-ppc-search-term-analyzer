"""
Step 2 validation: runs the analyzer, generates the 5-sheet Excel report
and the Amazon bulk negation CSV, then prints confirmation.
"""

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.analyzer import SearchTermAnalyzer
from app.exporter import generate_report
from app.bulk_generator import generate_bulk_csv
from tests.generate_dummy_data import generate

THRESHOLDS = {
    "click_threshold": 10,
    "acos_threshold": 0.30,
    "cvr_threshold": 0.10,
    "low_click_threshold": 10,
    "order_threshold": 2,
}

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")
SEPARATOR = "=" * 70


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    csv_path = os.path.join(os.path.dirname(__file__), "dummy_spstr.csv")
    if not os.path.exists(csv_path):
        generate(csv_path)

    analyzer = SearchTermAnalyzer()
    analyzer.load_data(csv_path)
    results = analyzer.analyze(THRESHOLDS)

    print(SEPARATOR)
    print("  STEP 2: EXPORT VALIDATION")
    print(SEPARATOR)

    # ---- Generate Excel Report ----
    report_path = os.path.join(OUTPUT_DIR, "Analysis_Report.xlsx")
    generate_report(results, output_path=report_path)
    report_size = os.path.getsize(report_path)
    print(f"\n  Excel Report: {report_path}")
    print(f"  Size: {report_size:,} bytes")

    import openpyxl
    wb = openpyxl.load_workbook(report_path)
    print(f"  Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"    {name}: {ws.max_row} rows x {ws.max_column} cols")
    wb.close()

    # ---- Generate Bulk Negation CSV ----
    wasted_df = results["Wasted Adspend"]
    bulk_path = os.path.join(OUTPUT_DIR, "Bulk_Negation.csv")
    generate_bulk_csv(wasted_df, output_path=bulk_path)
    bulk_size = os.path.getsize(bulk_path)
    print(f"\n  Bulk Negation CSV: {bulk_path}")
    print(f"  Size: {bulk_size:,} bytes")

    import pandas as pd
    bulk_df = pd.read_csv(bulk_path)
    print(f"  Rows: {len(bulk_df)}")
    print(f"  Columns: {list(bulk_df.columns)}")
    print(f"\n  Sample rows:")
    pd.set_option("display.max_colwidth", 40)
    pd.set_option("display.width", 200)
    print(bulk_df[["Entity", "Campaign Name (Informational only)", "Keyword Text", "Match Type", "State"]].to_string(index=False))

    # ---- Verify against reference ----
    ref_path = os.path.join(os.path.dirname(__file__), "reference_negation_bulk.xlsx")
    if os.path.exists(ref_path):
        ref_df = pd.read_excel(ref_path, engine="openpyxl")
        ref_cols = list(ref_df.columns)
        gen_cols = list(bulk_df.columns)
        match = ref_cols == gen_cols
        print(f"\n  Column order matches reference: {match}")
        if not match:
            print(f"    Reference: {ref_cols}")
            print(f"    Generated: {gen_cols}")

    # ---- Also test in-memory mode (for API streaming) ----
    report_bytes = generate_report(results)
    bulk_bytes = generate_bulk_csv(wasted_df)
    print(f"\n  In-memory report: {len(report_bytes):,} bytes")
    print(f"  In-memory bulk CSV: {len(bulk_bytes):,} bytes")

    print(f"\n{SEPARATOR}")
    print("  STEP 2 COMPLETE")
    print(SEPARATOR)


if __name__ == "__main__":
    main()
